from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[2]

INPUT = PROJECT_ROOT / "output" / "peer_percentiles.csv"
OUTPUT = PROJECT_ROOT / "output" / "peer_comparison.xlsx"

df = pd.read_csv(INPUT)

# Create Excel workbook
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:

    for peer_group in sorted(df["peer_group_name"].dropna().unique()):

        sheet = df[df["peer_group_name"] == peer_group]

        sheet.to_excel(
            writer,
            sheet_name=peer_group[:31],
            index=False
        )

# Apply colour formatting
wb = load_workbook(OUTPUT)

green = PatternFill(fill_type="solid", start_color="C6EFCE")
yellow = PatternFill(fill_type="solid", start_color="FFEB9C")
red = PatternFill(fill_type="solid", start_color="FFC7CE")

for ws in wb.worksheets:

    headers = [cell.value for cell in ws[1]]

    if "percentile_rank" not in headers:
        continue

    col = headers.index("percentile_rank") + 1

    for row in range(2, ws.max_row + 1):

        value = ws.cell(row=row, column=col).value

        if value is None:
            continue

        if value >= 75:
            ws.cell(row=row, column=col).fill = green

        elif value >= 50:
            ws.cell(row=row, column=col).fill = yellow

        else:
            ws.cell(row=row, column=col).fill = red

wb.save(OUTPUT)

print("=" * 60)
print("Peer Comparison Workbook Created")
print("=" * 60)
print(f"Sheets : {len(wb.sheetnames)}")
print(f"Output : {OUTPUT}")
